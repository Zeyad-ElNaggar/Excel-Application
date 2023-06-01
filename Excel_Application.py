#The excel file must be taken as a "copy path"and pasted in line 8 in the ("....")
#It is better that the copy path to have double slash(//) as it is shown in line 8


#from openpyxl import*
import openpyxl

wb = openpyxl.load_workbook("Countries (excel file).xlsx")
sheets = wb.sheetnames
print(sheets)

country_name = input("Please enter a country name : ")
while country_name not in sheets:
    country_name = input("Sorry there is no country with that name in the fie, Try again : ")

sheets = wb[country_name]

def population():

    row = sheets.max_row
    column = sheets.max_column

    for i in range(1, row + 1):
        for j in range(1, column + 1):
            print(sheets.cell(i, j).value)



def min_pop():

    list=[]
    second_column= sheets['B']

    for j in range(len(second_column)):
        list.append(second_column[j].value)

    del list[0]
    sorted_list=list.copy()
    sorted_list.sort()

    index=list.index(sorted_list[0])
    print("The lowest populaion governrate/state is ", sheets.cell(index+2,1).value ," ", sheets.cell(index+2,2).value )


def max_pop():
    list = []
    second_column = sheets['B']

    for j in range(len(second_column)):
        list.append(second_column[j].value)

    del list[0]
    sorted_list=list.copy()
    sorted_list.sort()

    index=list.index(sorted_list[-2])
    print("The highest populaion governrate/state is ", sheets.cell(index + 2, 1).value, " ", sheets.cell(index + 2, 2).value)
    input()


population()

x=input("Enter 'next' to display the max and min population city\n Enter 'exit' to exit :  ")
if x=='next':
    max_pop()
    min_pop()
else:
    exit()